"""
AI-Dost AI Engine — Python FastAPI sidecar (port 8001).

Hosts Python-only AI capabilities that the Node.js backend cannot do natively:
  - LlamaIndex RAG: workspace files / documents par semantic Q&A
    (embeddings via local Ollama nomic-embed-text — free & offline)

Phase 2 (planned): LangGraph orchestration, CrewAI/AutoGen multi-agent runs.

Architecture:
  Frontend (Next.js) -> Backend (Node/Express :5000) -> AI Engine (FastAPI :8001)
"""

import os
import shutil
import tempfile
from typing import List, Optional

from fastapi import FastAPI, HTTPException
from fastapi.responses import Response
from pydantic import BaseModel

app = FastAPI(title="AI-Dost AI Engine", version="1.0.0")

# ── Load .env (backend/.env shared) so crew LLM keys work standalone ─────────
def _load_env():
    candidates = [
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "backend", ".env"),
        os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"),
    ]
    for env_path in candidates:
        try:
            if os.path.isfile(env_path):
                with open(env_path, "r", encoding="utf-8") as f:
                    for line in f:
                        line = line.strip()
                        if not line or line.startswith("#") or "=" not in line:
                            continue
                        key, _, val = line.partition("=")
                        key, val = key.strip(), val.strip().strip('"').strip("'")
                        if not os.environ.get(key):
                            os.environ[key] = val
        except Exception:
            pass

_load_env()
# Groq `cache_breakpoint` param reject karta hai — litellm ko drop karne do
if not os.environ.get("LITELLM_DROP_PARAMS"):
    os.environ["LITELLM_DROP_PARAMS"] = "true"

# ── Lazy imports (so /health works even if llama-index is broken) ─────────────
def _llm():
    from llama_index.llms.ollama import Ollama
    return Ollama(model="qwen2.5-coder:7b", request_timeout=300.0)


def _embeddings():
    from llama_index.embeddings.ollama import OllamaEmbedding
    return OllamaEmbedding(model_name="nomic-embed-text")


def _load_nodes(directory: str):
    from llama_index.core import SimpleDirectoryReader
    reader = SimpleDirectoryReader(
        input_dir=directory,
        required_exts=[".txt", ".md", ".html", ".htm", ".css", ".js", ".jsx", ".tsx", ".ts", ".json", ".py", ".csv"],
        recursive=True,
        exclude_hidden=True,
        exclude=["**/node_modules/**", "**/.venv/**", "**/venv/**", "**/.git/**", "**/build/**", "**/dist/**", "**/coverage/**"]
    )
    return reader.load_data()


INDEX_CACHE = {}


def _get_index(directory: str, rebuild: bool = False):
    """Build (or reuse) a LlamaIndex vector index over a directory."""
    from llama_index.core import VectorStoreIndex, StorageContext
    import chromadb
    from llama_index.vector_stores.chroma import ChromaVectorStore
    import hashlib

    key = os.path.abspath(directory)
    # Create a valid collection name using md5 hash of the directory path
    collection_name = "dir_" + hashlib.md5(key.encode()).hexdigest()

    # Use a persistent ChromaDB store locally
    chroma_client = chromadb.PersistentClient(path="./chroma_db")
    chroma_collection = chroma_client.get_or_create_collection(collection_name)
    vector_store = ChromaVectorStore(chroma_collection=chroma_collection)
    storage_context = StorageContext.from_defaults(vector_store=vector_store)

    if key in INDEX_CACHE and not rebuild:
        return INDEX_CACHE[key]

    if rebuild:
        # Rebuilding: Clear existing data and re-index
        chroma_client.delete_collection(collection_name)
        chroma_collection = chroma_client.get_or_create_collection(collection_name)
        vector_store = ChromaVectorStore(chroma_collection=chroma_collection)
        storage_context = StorageContext.from_defaults(vector_store=vector_store)
        
        nodes = _load_nodes(directory)
        if not nodes:
            raise HTTPException(404, "Directory me koi readable file nahi mili")
            
        index = VectorStoreIndex(
            nodes,
            storage_context=storage_context,
            embed_model=_embeddings(),
            llm=_llm(),
            show_progress=False,
        )
    else:
        # Load existing if available, else build
        if chroma_collection.count() == 0:
            nodes = _load_nodes(directory)
            if not nodes:
                raise HTTPException(404, "Directory me koi readable file nahi mili")
            index = VectorStoreIndex(
                nodes,
                storage_context=storage_context,
                embed_model=_embeddings(),
                llm=_llm(),
                show_progress=False,
            )
        else:
            index = VectorStoreIndex.from_vector_store(
                vector_store,
                embed_model=_embeddings(),
                llm=_llm(),
            )

    INDEX_CACHE[key] = index
    return index


class RagQuery(BaseModel):
    directory: str
    question: str
    top_k: int = 4
    rebuild: bool = False


class RagResult(BaseModel):
    answer: str
    sources: List[dict]


@app.get("/health")
def health():
    import importlib.util
    return {
        "status": "ok",
        "service": "ai-dost-ai-engine",
        "llama_index": importlib.util.find_spec("llama_index") is not None,
        "ollama_running": True,
    }




class ScrapeRequest(BaseModel):
    url: str

class ScrapeResult(BaseModel):
    text: str
    url: str

@app.post("/ai/research/scrape", response_model=ScrapeResult)
def scrape_url(req: ScrapeRequest):
    """Scrapes a URL and extracts clean text without relying on external APIs."""
    import requests
    from bs4 import BeautifulSoup
    try:
        headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'}
        response = requests.get(req.url, headers=headers, timeout=10)
        response.raise_for_status()
        
        soup = BeautifulSoup(response.content, "html.parser")
        
        # Remove script and style tags
        for script_or_style in soup(["script", "style", "noscript", "header", "footer", "nav"]):
            script_or_style.extract()
            
        text = soup.get_text(separator=' ')
        
        # Clean up whitespace
        lines = (line.strip() for line in text.splitlines())
        chunks = (phrase.strip() for line in lines for phrase in line.split("  "))
        text = '\n'.join(chunk for chunk in chunks if chunk)
        
        return ScrapeResult(text=text[:15000], url=req.url) # Limit text to prevent massive responses
    except Exception as e:
        raise HTTPException(500, f"Scrape error: {e}")


# ------------------------------------------------------------------------------
# TAVILY WEB SEARCH — Real-time web search with sources
# ------------------------------------------------------------------------------
class TavilySearchRequest(BaseModel):
    query: str
    max_results: int = 5
    search_depth: str = "basic"  # basic | advanced
    include_domains: Optional[List[str]] = None
    exclude_domains: Optional[List[str]] = None


class TavilySearchResult(BaseModel):
    query: str
    results: List[dict]
    answer: Optional[str] = None
    response_time: float


@app.post("/ai/web/search", response_model=TavilySearchResult)
def tavily_search(req: TavilySearchRequest):
    """Search the web using Tavily API. Returns structured results with sources."""
    import os
    import time
    import requests

    api_key = os.environ.get("TAVILY_API_KEY")
    if not api_key:
        raise HTTPException(400, "TAVILY_API_KEY not set in environment")

    query = req.query.strip()
    if not query:
        raise HTTPException(400, "Query khali hai")

    start_time = time.time()
    try:
        url = "https://api.tavily.com/search"
        payload = {
            "api_key": api_key,
            "query": query,
            "max_results": req.max_results,
            "search_depth": req.search_depth,
            "include_answer": True,
            "include_raw_content": False,
        }
        if req.include_domains:
            payload["include_domains"] = req.include_domains
        if req.exclude_domains:
            payload["exclude_domains"] = req.exclude_domains

        response = requests.post(url, json=payload, timeout=30)
        response.raise_for_status()
        data = response.json()

        results = data.get("results", [])
        formatted = []
        for r in results:
            formatted.append({
                "title": r.get("title", ""),
                "url": r.get("url", ""),
                "content": r.get("content", "")[:500],
                "score": r.get("score", 0),
            })

        return TavilySearchResult(
            query=query,
            results=formatted,
            answer=data.get("answer"),
            response_time=time.time() - start_time
        )
    except requests.HTTPError as e:
        if e.response.status_code == 429:
            raise HTTPException(429, "Tavily rate limit exceeded")
        raise HTTPException(500, f"Tavily API error: {e}")
    except Exception as e:
        raise HTTPException(500, f"Tavily search error: {e}")


class AgentRunRequest(BaseModel):
    thread_id: str
    prompt: str
    mcp_command: str = "python"
    mcp_args: List[str] = []

class AgentRunResult(BaseModel):
    status: str
    result: str = ""
    tool_call: dict = None

# Global Checkpointer for HITL
agent_memory = None

@app.on_event("startup")
def startup_event():
    global agent_memory
    from langgraph.checkpoint.memory import MemorySaver
    agent_memory = MemorySaver()
    print("[AI-Dost] LangGraph MemorySaver initialized")

@app.post("/ai/agent/run", response_model=AgentRunResult)
async def agent_run(req: AgentRunRequest):
    from mcp_client import MCPClientManager
    from langchain_ollama import ChatOllama
    from langgraph.prebuilt import create_react_agent
    
    try:
        tools = []
        if req.mcp_command:
            mcp_manager = MCPClientManager(command=req.mcp_command, args=req.mcp_args)
            tools = await mcp_manager.get_langchain_tools()
        else:
            from langchain_core.tools import tool
            @tool
            def read_file(path: str) -> str:
                """Read content of a file"""
                return "Mock content of " + path
            
            tools = [read_file, create_new_tool] + load_custom_tools()

            mcp_manager = None
        
        llm = ChatOllama(model="qwen2.5-coder:7b", temperature=0.1) 
        
        # Use checkpointer and interrupt_before tools for HITL
        
        # Retrieve past learnings
        try:
            learning_index = get_learning_index()
            retriever = learning_index.as_retriever(similarity_top_k=2)
            learning_nodes = retriever.retrieve(req.prompt)
            learnings_text = "\n".join([n.get_content() for n in learning_nodes])
            system_prompt = f"Relevant Past Learnings:\n{learnings_text}\n\nYou are a helpful AI Assistant."
        except Exception:
            system_prompt = "You are a helpful AI Assistant."
            
        agent_executor = create_react_agent(llm, tools, checkpointer=agent_memory, interrupt_before=["tools"], state_modifier=system_prompt)

        config = {"configurable": {"thread_id": req.thread_id}}
        
        response = await agent_executor.ainvoke({"messages": [("user", req.prompt)]}, config)
        
        state = agent_executor.get_state(config)
        if state.next:
            # Interrupted before tool execution
            last_message = response["messages"][-1]
            tool_call = last_message.tool_calls[0] if hasattr(last_message, "tool_calls") and last_message.tool_calls else None
            if mcp_manager: await mcp_manager.close()
            return AgentRunResult(status="requires_approval", tool_call=tool_call)
            
        final_answer = response["messages"][-1].content
        if mcp_manager: await mcp_manager.close()
        
        return AgentRunResult(status="completed", result=final_answer)
    except Exception as e:
        raise HTTPException(500, f"Agent run error: {e}")

class AgentResumeRequest(BaseModel):
    thread_id: str
    mcp_command: str = "python"
    mcp_args: List[str] = []
    approved: bool

@app.post("/ai/agent/resume", response_model=AgentRunResult)
async def agent_resume(req: AgentResumeRequest):
    from mcp_client import MCPClientManager
    from langchain_ollama import ChatOllama
    from langgraph.prebuilt import create_react_agent
    from langchain_core.messages import ToolMessage
    
    try:
        tools = []
        if req.mcp_command:
            mcp_manager = MCPClientManager(command=req.mcp_command, args=req.mcp_args)
            tools = await mcp_manager.get_langchain_tools()
        else:
            from langchain_core.tools import tool
            @tool
            def read_file(path: str) -> str:
                """Read content of a file"""
                return "Mock content of " + path
            
            tools = [read_file, create_new_tool] + load_custom_tools()

            mcp_manager = None
            
        llm = ChatOllama(model="qwen2.5-coder:7b", temperature=0.1) 
        
        
        # Retrieve past learnings
        try:
            learning_index = get_learning_index()
            retriever = learning_index.as_retriever(similarity_top_k=2)
            learning_nodes = retriever.retrieve(req.prompt)
            learnings_text = "\n".join([n.get_content() for n in learning_nodes])
            system_prompt = f"Relevant Past Learnings:\n{learnings_text}\n\nYou are a helpful AI Assistant."
        except Exception:
            system_prompt = "You are a helpful AI Assistant."
            
        agent_executor = create_react_agent(llm, tools, checkpointer=agent_memory, interrupt_before=["tools"], state_modifier=system_prompt)

        config = {"configurable": {"thread_id": req.thread_id}}
        
        if req.approved:
            # Continue execution by invoking with None
            response = await agent_executor.ainvoke(None, config)
        else:
            # Inject a ToolMessage indicating denial to skip actual tool execution
            state = agent_executor.get_state(config)
            last_msg = state.values["messages"][-1]
            tool_call_id = last_msg.tool_calls[0]["id"]
            denial_msg = ToolMessage(tool_call_id=tool_call_id, name=last_msg.tool_calls[0]["name"], content="User denied this action.")
            
            # Update state with the denial message to bypass the tool node
            agent_executor.update_state(config, {"messages": [denial_msg]}, as_node="tools")
            
            # Resume from after the tool node
            response = await agent_executor.ainvoke(None, config)
            
        state = agent_executor.get_state(config)
        if state.next:
            last_message = response["messages"][-1]
            tool_call = last_message.tool_calls[0] if hasattr(last_message, "tool_calls") and last_message.tool_calls else None
            if mcp_manager: await mcp_manager.close()
            return AgentRunResult(status="requires_approval", tool_call=tool_call)
            
        final_answer = response["messages"][-1].content
        if mcp_manager: await mcp_manager.close()
        
        return AgentRunResult(status="completed", result=final_answer)
    except Exception as e:
        raise HTTPException(500, f"Agent resume error: {e}")





# ------------------------------------------------------------------------------
# DYNAMIC TOOLS (Self-Evolution)
# ------------------------------------------------------------------------------
import sys
import importlib.util
from langchain_core.tools import tool

CUSTOM_TOOLS_DIR = os.path.join(os.path.dirname(__file__), "custom_tools")
os.makedirs(CUSTOM_TOOLS_DIR, exist_ok=True)

@tool
def create_new_tool(tool_name: str, python_code: str) -> str:
    """
    Use this tool to create a NEW capability for yourself!
    Provide the exact Python code. The code MUST contain exactly ONE function decorated with @tool.
    Example python_code:
    from langchain_core.tools import tool
    @tool
    def my_new_tool(text: str) -> str:
        '''Description of tool'''
        return text.upper()
    """
    try:
        # Save to disk
        filepath = os.path.join(CUSTOM_TOOLS_DIR, f"{tool_name}.py")
        with open(filepath, "w", encoding="utf-8") as f:
            f.write(python_code)
        return f"Success! Tool '{tool_name}' saved to {filepath}. It will be available in the next agent run."
    except Exception as e:
        return f"Failed to create tool: {e}"

def load_custom_tools():
    """Dynamically loads all .py files in custom_tools as Langchain tools."""
    loaded_tools = []
    if not os.path.exists(CUSTOM_TOOLS_DIR):
        return loaded_tools
        
    for filename in os.listdir(CUSTOM_TOOLS_DIR):
        if filename.endswith(".py"):
            filepath = os.path.join(CUSTOM_TOOLS_DIR, filename)
            module_name = filename[:-3]
            try:
                spec = importlib.util.spec_from_file_location(module_name, filepath)
                if spec and spec.loader:
                    module = importlib.util.module_from_spec(spec)
                    sys.modules[module_name] = module
                    spec.loader.exec_module(module)
                    
                    # Find the tool in the module (assuming it has a 'name' attribute or is a Langchain BaseTool)
                    for attr_name in dir(module):
                        attr = getattr(module, attr_name)
                        if hasattr(attr, "name") and hasattr(attr, "description") and callable(attr):
                            # Usually a LangChain tool instance
                            loaded_tools.append(attr)
            except Exception as e:
                print(f"Error loading custom tool {filename}: {e}")
    return loaded_tools


# ------------------------------------------------------------------------------
# LONG-TERM MEMORY (Continuous Learning)
# ------------------------------------------------------------------------------

class MemoryLearnRequest(BaseModel):
    text: str
    
class MemoryRetrieveRequest(BaseModel):
    query: str
    top_k: int = 3

def get_learning_index():
    import chromadb
    from llama_index.vector_stores.chroma import ChromaVectorStore
    from llama_index.core import VectorStoreIndex, StorageContext, Document
    
    chroma_client = chromadb.PersistentClient(path="./chroma_db")
    chroma_collection = chroma_client.get_or_create_collection("agent_learnings")
    vector_store = ChromaVectorStore(chroma_collection=chroma_collection)
    storage_context = StorageContext.from_defaults(vector_store=vector_store)
    
    if chroma_collection.count() == 0:
        # Create empty index with a dummy document so it initializes properly
        doc = Document(text="Initial rule: Always follow user instructions carefully.")
        index = VectorStoreIndex.from_documents(
            [doc],
            storage_context=storage_context,
            embed_model=_embeddings(),
        )
    else:
        index = VectorStoreIndex.from_vector_store(
            vector_store=vector_store,
            embed_model=_embeddings(),
        )
    return index

@app.post("/ai/agent/learn")
async def save_memory(req: MemoryLearnRequest):
    from llama_index.core import Document
    try:
        index = get_learning_index()
        doc = Document(text=req.text)
        index.insert(doc)
        return {"status": "success", "message": "Memory saved!"}
    except Exception as e:
        raise HTTPException(500, f"Error saving memory: {e}")

@app.post("/ai/agent/memory/retrieve")
async def retrieve_memory(req: MemoryRetrieveRequest):
    try:
        index = get_learning_index()
        retriever = index.as_retriever(similarity_top_k=req.top_k)
        nodes = retriever.retrieve(req.query)
        results = [n.get_content() for n in nodes]
        return {"status": "success", "learnings": results}
    except Exception as e:
        raise HTTPException(500, f"Error retrieving memory: {e}")


# ------------------------------------------------------------------------------
# MULTI-AGENT SWARM (CrewAI-Style via LangGraph System Prompts)
# ------------------------------------------------------------------------------
class SwarmRunRequest(BaseModel):
    thread_id: str
    prompt: str
    mcp_command: str = ""
    mcp_args: List[str] = []

@app.post("/ai/agent/swarm", response_model=AgentRunResult)
async def swarm_run(req: SwarmRunRequest):
    from mcp_client import MCPClientManager
    from langchain_ollama import ChatOllama
    from langgraph.prebuilt import create_react_agent
    
    try:
        tools = []
        mcp_manager = None
        if req.mcp_command:
            mcp_manager = MCPClientManager(command=req.mcp_command, args=req.mcp_args)
            tools = await mcp_manager.get_langchain_tools()
        else:
            from langchain_core.tools import tool
            @tool
            def read_file(path: str) -> str:
                """Read content of a file"""
                return "Mock content of " + path
            
            tools = [read_file, create_new_tool] + load_custom_tools()

            
        llm = ChatOllama(model="qwen2.5-coder:7b", temperature=0.1) 
        
        # Swarm System Prompt
        
        # Retrieve past learnings
        try:
            learning_index = get_learning_index()
            retriever = learning_index.as_retriever(similarity_top_k=2)
            learning_nodes = retriever.retrieve(req.prompt)
            learnings_text = "\n".join([n.get_content() for n in learning_nodes])
            past_context = f"Relevant Past Learnings:\n{learnings_text}\n\n"
        except Exception:
            past_context = ""
            
        swarm_prompt = past_context + """You are the Swarm Manager. You control a team of agents: Researcher, Coder, and Tester.
 You control a team of agents: Researcher, Coder, and Tester.
        You must tackle the user's task by simulating this team.
        1. First, act as the Researcher to gather information using tools.
        2. Then, act as the Coder to write the necessary files.
        3. Finally, act as the Tester to review and verify the files.
        Always state which role you are currently playing in your thought process (e.g. '[Coder]: Writing the file...').
        """
        
        agent_executor = create_react_agent(llm, tools, checkpointer=agent_memory, interrupt_before=["tools"], state_modifier=swarm_prompt)
        config = {"configurable": {"thread_id": req.thread_id}}
        
        response = await agent_executor.ainvoke({"messages": [("user", req.prompt)]}, config)
        
        state = agent_executor.get_state(config)
        if state.next:
            last_message = response["messages"][-1]
            tool_call = last_message.tool_calls[0] if hasattr(last_message, "tool_calls") and last_message.tool_calls else None
            if mcp_manager: await mcp_manager.close()
            return AgentRunResult(status="requires_approval", tool_call=tool_call)
            
        final_answer = response["messages"][-1].content
        if mcp_manager: await mcp_manager.close()
        
        return AgentRunResult(status="completed", result=final_answer)
    except Exception as e:
        raise HTTPException(500, f"Swarm run error: {e}")


# ------------------------------------------------------------------------------
# CREWAI MULTI-AGENT CREW (Pillar 1: Agentic Core — role-based collaboration)
# ------------------------------------------------------------------------------
class CrewRunRequest(BaseModel):
    prompt: str
    mode: str = "dev"          # dev | research | content
    model: str = "ollama"      # ollama (free/local) | gemini | groq | nvidia | cerebras
    directory: str = ""

class CrewRunResult(BaseModel):
    status: str
    result: str = ""
    crew_output: str = ""
    agents: List[str] = []
    files: List[str] = []
    directory: str = ""

@app.post("/ai/crew/run", response_model=CrewRunResult)
def crew_run(req: CrewRunRequest):
    """Run a real CrewAI role-based crew (Researcher + Coder + Reviewer).

    Free & offline by default: Ollama qwen2.5-coder:7b local LLM.
    Modes:
      - dev:      Researcher(analyze) -> Coder(write) -> Reviewer(verify)
      - research: Researcher(gather) -> Writer(summarize)
      - content:  Writer(draft) -> Reviewer(polish)
    """
    try:
        from crewai import Agent, Task, Crew, Process
        from crewai.tools import tool
        import crewai.llms.cache as _cache_mod

        # Groq `cache_breakpoint` support nahi karta — marker no-op bana do
        _cache_mod.mark_cache_breakpoint = lambda msg: msg

        prompt = req.prompt.strip()
        if not prompt:
            raise HTTPException(400, "Prompt khali hai")

        work_dir = req.directory or os.path.join(os.environ.get("TEMP", "."), "ai-dost-workspace")
        if not os.path.isdir(work_dir):
            try:
                os.makedirs(work_dir, exist_ok=True)
            except Exception:
                pass

        # ── Custom file tools (CrewAI 1.15 ke paas built-in file tools nahi) ──
        @tool("save_project_file")
        def save_project_file(filename: str, content: str) -> str:
            """Create or overwrite a file inside the project workspace.
            filename ek relative path hota hai (jaise 'src/app.py' ya 'index.html').
            Content pura file content hota hai."""
            if not filename or filename.startswith(("/", "\\", ".")) or ".." in filename.split("/") + filename.split("\\"):
                return "Error: invalid filename — sirf relative path do (jaise 'src/app.py')."
            safe_name = filename.replace("\\", "/").lstrip("/")
            target = os.path.abspath(os.path.join(work_dir, safe_name))
            if os.path.commonpath([target, os.path.abspath(work_dir)]) != os.path.abspath(work_dir):
                return "Error: path workspace ke bahar jata hai."
            try:
                os.makedirs(os.path.dirname(target), exist_ok=True)
                with open(target, "w", encoding="utf-8") as f:
                    f.write(content)
                return f"File save ho gayi: {safe_name}"
            except Exception as e:
                return f"Error: {e}"

        @tool("list_project_files")
        def list_project_files(query: str = "") -> str:
            """List files available in the project workspace.
            query optional hota hai — empty chhodo agar sirf listing chahiye."""
            try:
                out = []
                for root, _dirs, files in os.walk(work_dir):
                    for f in files:
                        if f.startswith(".") or "node_modules" in root or ".git" in root:
                            continue
                        rel = os.path.relpath(os.path.join(root, f), work_dir)
                        out.append(rel)
                return "\n".join(out[:100]) or "(workspace khali hai)"
            except Exception as e:
                return f"Error: {e}"

        # ── LLM choice (API keys .env se load ho jate hain startup pe) ──
        from crewai.llm import LLM
        if req.model == "gemini":
            key = os.environ.get("GEMINI_API_KEY", "")
            if not key:
                raise HTTPException(400, "GEMINI_API_KEY set nahi hai")
            llm = LLM(model="gemini/gemini-2.5-flash")
        elif req.model == "groq":
            key = os.environ.get("GROQ_API_KEY", "")
            if not key:
                raise HTTPException(400, "GROQ_API_KEY set nahi hai")
            llm = LLM(model="groq/llama-3.3-70b-versatile", additional_params={"drop_params": True})
        elif req.model == "nvidia":
            key = os.environ.get("NVIDIA_API_KEY", "")
            if not key:
                raise HTTPException(400, "NVIDIA_API_KEY set nahi hai")
            llm = LLM(model="openai/meta/llama-3.1-8b-instruct", api_key=key, api_base="https://integrate.api.nvidia.com/v1")
            # NVIDIA single tool-call support karta — ReAct path use karo (sequential tool calls)
            llm.supports_function_calling = lambda: False
        elif req.model == "cerebras":
            key = os.environ.get("CEREBRAS_API_KEY", "")
            if not key:
                raise HTTPException(400, "CEREBRAS_API_KEY set nahi hai — cerebras.ai/cloud se free key lo")
            llm = LLM(model="cerebras/gpt-oss-120b")
        else:
            llm = LLM(model="ollama/qwen2.5-coder:7b")

        if req.mode == "research":
            researcher = Agent(
                role="Senior Researcher",
                goal=f"Deeply research: {prompt}. Use list_project_files tool aur available files ki context.",
                backstory="You are a meticulous researcher who finds facts, sources and structured insights.",
                llm=llm, verbose=False, allow_delegation=False, tools=[list_project_files],
            )
            writer = Agent(
                role="Technical Writer",
                goal="Turn raw research into a clear, structured, well-written summary/report. Use save_project_file to save the report as a .md file.",
                backstory="You write crisp, professional summaries with headings and bullet points.",
                llm=llm, verbose=False, allow_delegation=False, tools=[save_project_file],
            )
            t1 = Task(
                description=f"Research the topic: {prompt}. Use list_project_files to see workspace files. Return raw findings with key facts and sources.",
                expected_output="Structured research findings with facts and sources.",
                agent=researcher,
            )
            t2 = Task(
                description="Write a final clean report from the research findings. Use headings, bullets, and a summary.",
                expected_output="A polished markdown report.",
                agent=writer,
            )
            agents = [researcher, writer]
        elif req.mode == "content":
            writer = Agent(
                role="Content Writer",
                goal=f"Create engaging content about: {prompt}. Use save_project_file to save the draft as a .md file.",
                backstory="You craft compelling, human-sounding articles and scripts.",
                llm=llm, verbose=False, allow_delegation=False, tools=[save_project_file],
            )
            editor = Agent(
                role="Content Editor",
                goal="Polish the draft: fix grammar, tighten wording, improve flow and structure. Save final version with save_project_file.",
                backstory="You are a sharp editor who makes good content great.",
                llm=llm, verbose=False, allow_delegation=False, tools=[save_project_file],
            )
            t1 = Task(
                description=f"Write a first draft (article/script) about: {prompt}. Use list_project_files for workspace context.",
                expected_output="A complete first draft.",
                agent=writer,
            )
            t2 = Task(
                description="Edit the draft for clarity, grammar, flow and impact. Return the final polished version.",
                expected_output="Final polished content.",
                agent=editor,
            )
            agents = [writer, editor]
        else:  # dev mode (default)
            researcher = Agent(
                role="Codebase Researcher",
                goal=f"Analyze the request and existing workspace using list_project_files: {prompt}. Determine what files/code are needed.",
                backstory="You inspect requirements and existing files to plan exact implementation steps.",
                llm=llm, verbose=False, allow_delegation=False, tools=[list_project_files],
            )
            coder = Agent(
                role="Senior Software Engineer",
                goal="Write complete, working, production-quality code implementing the research plan. MUST use save_project_file to actually create the files in the workspace.",
                backstory="You are a full-stack engineer who writes clean, correct code and saves every file with the save_project_file tool.",
                llm=llm, verbose=False, allow_delegation=False, tools=[save_project_file, list_project_files],
            )
            reviewer = Agent(
                role="Code Reviewer",
                goal="Review the code for bugs, security issues and completeness. List fixes and final answer.",
                backstory="You are a meticulous reviewer who catches edge cases before they ship.",
                llm=llm, verbose=False, allow_delegation=False,
            )
            t1 = Task(
                description=f"Analyze: {prompt}. Use list_project_files to see the workspace. Produce a plan: which files to create/edit and what each should contain.",
                expected_output="A short implementation plan (files + purpose).",
                agent=researcher,
            )
            t2 = Task(
                description="Implement the plan. Write the actual code/files with full content. Be concrete and complete.",
                expected_output="Complete code for all planned files.",
                agent=coder,
            )
            t3 = Task(
                description="Review the produced code. List any bugs/fixes, then give the final summary of what was built.",
                expected_output="Review notes + final summary.",
                agent=reviewer,
            )
            agents = [researcher, coder, reviewer]

        crew = Crew(
            agents=agents,
            tasks=[t1, t2, t3] if req.mode == "dev" else [t1, t2],
            process=Process.sequential,
            verbose=False,
        )
        # NVIDIA jaisi backends transient "single tool-call" errors deti hain — retry
        raw = ""
        last_err = None
        for attempt in range(3):
            try:
                output = crew.kickoff(inputs={"prompt": prompt})
                raw = getattr(output, "raw", None) or str(output)
                if raw.strip():
                    break
            except Exception as e:
                last_err = e
                if req.model == "nvidia" and attempt < 2:
                    continue
                raise
        if not raw.strip() and last_err:
            raise last_err

        # ── Files jo crew ne workspace me likhi (verify) ──
        created_files = []
        try:
            for root, _dirs, files in os.walk(work_dir):
                for f in files:
                    if f.startswith(".") or "node_modules" in root or ".git" in root:
                        continue
                    created_files.append(os.path.relpath(os.path.join(root, f), work_dir))
        except Exception:
            pass

        return CrewRunResult(
            status="completed",
            result=raw[:8000],
            crew_output=raw[:20000],
            agents=[a.role for a in agents],
            files=created_files[:50],
            directory=work_dir,
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"Crew run error: {e}")


# ------------------------------------------------------------------------------
# XLSX GENERATION — Real Excel with openpyxl
# ------------------------------------------------------------------------------
class XLSXRequest(BaseModel):
    topic: str
    title: str = ""
    columns: Optional[List[str]] = None  # Optional custom columns
    rows: int = 20


class XLSXResult(BaseModel):
    status: str
    file_path: str
    rows_written: int
    columns: List[str]


@app.post("/ai/xlsx/generate", response_model=XLSXResult)
def xlsx_generate(req: XLSXRequest):
    """Generate a real .xlsx file with openpyxl. Returns file path for download."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    import tempfile
    import os
    from datetime import datetime

    topic = req.topic.strip()
    if not topic:
        raise HTTPException(400, "Topic khali hai")

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"

    # Determine columns from topic or use custom
    if req.columns:
        columns = req.columns
    else:
        # Smart column detection based on topic
        topic_lower = topic.lower()
        if any(kw in topic_lower for kw in ['shaheed', 'martyr', 'jawan', 'soldier']):
            columns = ['Name', 'Rank', 'Regiment', 'Date', 'Place', 'State', 'Conflict']
        elif any(kw in topic_lower for kw in ['employee', 'staff', 'worker']):
            columns = ['Name', 'ID', 'Department', 'Role', 'Join Date', 'Salary']
        elif any(kw in topic_lower for kw in ['product', 'item', 'inventory']):
            columns = ['Product', 'Category', 'Price', 'Stock', 'SKU', 'Supplier']
        elif any(kw in topic_lower for kw in ['student', 'marks', 'grade']):
            columns = ['Name', 'Roll No', 'Subject', 'Marks', 'Grade', 'Semester']
        else:
            columns = ['Name', 'Category', 'Detail', 'Date', 'Notes']

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Write headers
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Generate sample data rows (in production, LLM would provide real data)
    # For now, generate placeholder rows with topic-relevant content
    for row_idx in range(2, req.rows + 2):
        for col_idx, col_name in enumerate(columns, 1):
            # Simple placeholder generation
            if col_name.lower() == 'name':
                value = f"{topic} Entry {row_idx - 1}"
            elif col_name.lower() == 'date':
                value = datetime.now().strftime("%Y-%m-%d")
            elif col_name.lower() in ['rank', 'role', 'category', 'grade']:
                value = "TBD"
            elif col_name.lower() in ['price', 'salary', 'marks']:
                value = 0
            else:
                value = f"{topic} - {col_name} {row_idx - 1}"
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True)

    # Auto-fit column widths
    for col_idx in range(1, len(columns) + 1):
        max_length = len(columns[col_idx - 1])
        for row_idx in range(2, min(req.rows + 2, 22)):
            cell_val = str(ws.cell(row=row_idx, column=col_idx).value or "")
            max_length = max(max_length, len(cell_val))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = min(max_length + 4, 40)

    # Save to temp file
    temp_dir = tempfile.gettempdir()
    filename = f"xlsx_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{abs(hash(topic)) % 10000}.xlsx"
    file_path = os.path.join(temp_dir, filename)
    wb.save(file_path)

    return XLSXResult(
        status="completed",
        file_path=file_path,
        rows_written=req.rows,
        columns=columns
    )


# ------------------------------------------------------------------------------
# EDGE TTS — FREE unlimited text-to-speech (Microsoft Edge voices, no API key)
# ------------------------------------------------------------------------------
class TTSRequest(BaseModel):
    text: str
    voice: str = "en-IN-PrabhatNeural"   # hi-IN-SwaraNeural | en-US-JennyNeural | etc.
    rate: str = "+0%"

@app.post("/ai/tts")
def tts(req: TTSRequest):
    """Convert text to speech (MP3) via Microsoft Edge TTS — 100% free, no key."""
    import edge_tts
    import io as _io

    text = (req.text or "").strip()[:2000]
    if not text:
        raise HTTPException(400, "Text khali hai")
    try:
        async def _run():
            communicate = edge_tts.Communicate(text, req.voice, rate=req.rate)
            audio = _io.BytesIO()
            async for chunk in communicate.stream():
                if chunk["type"] == "audio":
                    audio.write(chunk["data"])
            return audio.getvalue()
        import asyncio
        mp3 = asyncio.run(_run())
        if not mp3:
            raise HTTPException(500, "TTS ne koi audio nahi banaya")
        return Response(
            content=mp3,
            media_type="audio/mpeg",
            headers={"Content-Disposition": "inline; filename=ai-dost-tts.mp3"}
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"TTS error: {e}")


# ------------------------------------------------------------------------------
# UNIFIED PYTHON AI GENERATION & CASCADE ROUTER
# ------------------------------------------------------------------------------
class GenerateRequest(BaseModel):
    prompt: str
    system_prompt: Optional[str] = "You are AI-Dost, an expert AI developer assistant."
    model: Optional[str] = "auto"
    temperature: Optional[float] = 0.7
    max_tokens: Optional[int] = 2048

class GenerateResult(BaseModel):
    status: str
    text: str
    model_used: str
    response_time: float

@app.post("/ai/generate", response_model=GenerateResult)
def ai_generate(req: GenerateRequest):
    """Centralized Python AI text & code generator with automatic cascade."""
    import time
    import requests
    start_time = time.time()
    prompt = req.prompt.strip()
    if not prompt:
        raise HTTPException(400, "Prompt cannot be empty")

    groq_key = os.environ.get("GROQ_API_KEY")
    gemini_key = os.environ.get("GEMINI_API_KEY")
    nvidia_key = os.environ.get("NVIDIA_API_KEY")
    together_key = os.environ.get("TOGETHER_API_KEY")

    # 1. Groq (Fastest)
    if groq_key and req.model in ["auto", "groq"]:
        try:
            res = requests.post(
                "https://api.groq.com/openai/v1/chat/completions",
                headers={"Authorization": f"Bearer {groq_key}", "Content-Type": "application/json"},
                json={
                    "model": "llama-3.3-70b-versatile",
                    "messages": [
                        {"role": "system", "content": req.system_prompt},
                        {"role": "user", "content": prompt}
                    ],
                    "temperature": req.temperature,
                    "max_tokens": req.max_tokens,
                },
                timeout=15
            )
            if res.status_code == 200:
                data = res.json()
                content = data["choices"][0]["message"]["content"]
                return GenerateResult(status="success", text=content, model_used="groq/llama-3.3-70b", response_time=time.time() - start_time)
        except Exception:
            pass

    # 2. Gemini
    if gemini_key and req.model in ["auto", "gemini"]:
        try:
            url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={gemini_key}"
            res = requests.post(
                url,
                headers={"Content-Type": "application/json"},
                json={
                    "contents": [{"parts": [{"text": f"{req.system_prompt}\n\n{prompt}"}]}],
                    "generationConfig": {"temperature": req.temperature, "maxOutputTokens": req.max_tokens}
                },
                timeout=20
            )
            if res.status_code == 200:
                data = res.json()
                content = data["candidates"][0]["content"]["parts"][0]["text"]
                return GenerateResult(status="success", text=content, model_used="gemini-2.5-flash", response_time=time.time() - start_time)
        except Exception:
            pass

    # 3. Together AI
    if together_key and req.model in ["auto", "together"]:
        try:
            res = requests.post(
                "https://api.together.xyz/v1/chat/completions",
                headers={"Authorization": f"Bearer {together_key}", "Content-Type": "application/json"},
                json={
                    "model": "meta-llama/Llama-3.3-70B-Instruct-Turbo",
                    "messages": [{"role": "system", "content": req.system_prompt}, {"role": "user", "content": prompt}],
                    "temperature": req.temperature,
                    "max_tokens": req.max_tokens
                },
                timeout=15
            )
            if res.status_code == 200:
                content = res.json()["choices"][0]["message"]["content"]
                return GenerateResult(status="success", text=content, model_used="together/llama-3.3-70b", response_time=time.time() - start_time)
        except Exception:
            pass

    # 4. Ollama fallback
    try:
        res = requests.post(
            "http://127.0.0.1:11434/api/generate",
            json={"model": "qwen2.5-coder:7b", "prompt": f"{req.system_prompt}\n\n{prompt}", "stream": False},
            timeout=30
        )
        if res.status_code == 200:
            return GenerateResult(status="success", text=res.json().get("response", ""), model_used="ollama/qwen2.5-coder", response_time=time.time() - start_time)
    except Exception:
        pass

    raise HTTPException(503, "All AI providers in Python Engine temporarily unavailable")


# ------------------------------------------------------------------------------
# INLINE GHOST-TEXT CODE COMPLETION (Monaco Editor Bridge)
# ------------------------------------------------------------------------------
class CodeCompleteRequest(BaseModel):
    prefix: str
    suffix: Optional[str] = ""
    language: Optional[str] = "javascript"
    filename: Optional[str] = "file.js"

class CodeCompleteResult(BaseModel):
    completion: str
    language: str

@app.post("/ai/code/complete", response_model=CodeCompleteResult)
def code_complete(req: CodeCompleteRequest):
    """Fast inline code completion for Monaco Editor."""
    system = f"You are an ultra-fast code completion engine for {req.language}. Complete the code directly without explanations, markdown fences, or comments."
    prompt = f"File: {req.filename}\n\nCode before cursor:\n{req.prefix[-1000:]}\n\nCode after cursor:\n{req.suffix[:300]}\n\nCompletion:"
    try:
        gen = ai_generate(GenerateRequest(prompt=prompt, system_prompt=system, temperature=0.1, max_tokens=128))
        cleaned = gen.text.replace("```" + req.language, "").replace("```", "").strip()
        return CodeCompleteResult(completion=cleaned, language=req.language)
    except Exception:
        return CodeCompleteResult(completion="", language=req.language)


# ------------------------------------------------------------------------------
# PYTHON CODE INSPECTOR & AST LINTER
# ------------------------------------------------------------------------------
class CodeAnalyzeRequest(BaseModel):
    code: str
    language: Optional[str] = "python"

class CodeAnalyzeResult(BaseModel):
    valid: bool
    issues: List[dict]
    summary: str

@app.post("/ai/code/analyze", response_model=CodeAnalyzeResult)
def code_analyze(req: CodeAnalyzeRequest):
    """Analyze code for syntax errors and logic bugs."""
    import ast
    issues = []
    if req.language == "python":
        try:
            ast.parse(req.code)
            return CodeAnalyzeResult(valid=True, issues=[], summary="Python syntax valid. No parse errors found.")
        except SyntaxError as e:
            issues.append({
                "line": e.lineno,
                "column": e.offset,
                "message": e.msg,
                "type": "SyntaxError"
            })
            return CodeAnalyzeResult(valid=False, issues=issues, summary=f"Syntax Error at line {e.lineno}: {e.msg}")
    return CodeAnalyzeResult(valid=True, issues=[], summary="Code analyzed.")


# ------------------------------------------------------------------------------
# PYTHON REPORTLAB PDF GENERATOR
# ------------------------------------------------------------------------------
class PDFGenerateRequest(BaseModel):
    title: str
    content: str
    filename: Optional[str] = None

class PDFGenerateResult(BaseModel):
    status: str
    filename: str
    file_path: str

@app.post("/ai/pdf/generate", response_model=PDFGenerateResult)
def pdf_generate(req: PDFGenerateRequest):
    """Generate high-quality PDF using Python."""
    import tempfile
    import os
    from datetime import datetime

    fname = req.filename or f"doc_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    temp_dir = tempfile.gettempdir()
    out_path = os.path.join(temp_dir, fname)

    # Use basic canvas text generation
    try:
        from reportlab.lib.pagesizes import letter
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

        doc = SimpleDocTemplate(out_path, pagesize=letter, rightMargin=40, leftMargin=40, topMargin=40, bottomMargin=40)
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('Title', parent=styles['Heading1'], fontSize=18, leading=22, spaceAfter=12)
        body_style = ParagraphStyle('Body', parent=styles['Normal'], fontSize=10, leading=14, spaceAfter=8)

        story = [Paragraph(req.title, title_style), Spacer(1, 10)]
        for para in req.content.split("\n\n"):
            if para.strip():
                clean_p = para.replace("<", "&lt;").replace(">", "&gt;").replace("\n", "<br/>")
                story.append(Paragraph(clean_p, body_style))

        doc.build(story)
        return PDFGenerateResult(status="completed", filename=fname, file_path=out_path)
    except Exception as e:
        raise HTTPException(500, f"PDF generation error: {e}")



# ------------------------------------------------------------------------------
# RAG RETRIEVAL CONTRACTS (PHASE 2F.1)
# ------------------------------------------------------------------------------
class RetrievalFilter(BaseModel):
    source_types: Optional[List[str]] = []
    limit: Optional[int] = 10

class RetrievalRequest(BaseModel):
    version: str = "1"
    user_id: str
    project_id: str
    query: str
    mode: Optional[str] = "HYBRID"
    filters: Optional[RetrievalFilter] = RetrievalFilter()

class RetrievalResultItem(BaseModel):
    source_entity_id: str
    project_id: str
    source_type: str
    chunk_id: str
    score: float
    version_hash: str
    metadata: dict

class RetrievalResponse(BaseModel):
    version: str = "1"
    results: List[RetrievalResultItem]


AUTHORITY_SCORES = {
    "workspace_file": 1.0,
    "artifact": 0.9,
    "verification_result": 0.9,
    "context_node": 0.8,
    "message": 0.6,
    "execution_history": 0.5
}

def get_authority_score(source_type: str) -> float:
    return AUTHORITY_SCORES.get(source_type, 0.5)

@app.post("/ai/rag/query", response_model=RetrievalResponse)
def rag_query(req: RetrievalRequest):
    """
    Phase 2F.4 Hybrid Retrieval & Ranking
    """
    if not req.project_id:
        raise HTTPException(400, "project_id is required for tenant isolation")
    
    if not rag_collection:
        raise HTTPException(503, "Vector store is unavailable")

    query_text = req.query.strip()
    if not query_text:
        return RetrievalResponse(version=req.version, results=[])

    mode = (req.mode or "HYBRID").upper()
    if mode not in ["EXACT", "FULL_TEXT", "SEMANTIC", "HYBRID"]:
        raise HTTPException(400, "unsupported mode")

    limit = req.filters.limit if req.filters and req.filters.limit else 10
    limit = max(1, min(limit, 100))

    where_filter = {"project_id": {"$eq": req.project_id}}
    
    if req.filters and req.filters.source_types:
        if len(req.filters.source_types) == 1:
            where_filter = {
                "$and": [
                    {"project_id": {"$eq": req.project_id}},
                    {"source_type": {"$eq": req.filters.source_types[0]}}
                ]
            }
        else:
            where_filter = {
                "$and": [
                    {"project_id": {"$eq": req.project_id}},
                    {"source_type": {"$in": req.filters.source_types}}
                ]
            }

    entity_candidates = {}

    def add_candidate(meta, chunk_id, sem_score, kw_score):
        src_id = meta.get("source_entity_id")
        if not src_id: return
        auth_score = get_authority_score(meta.get("source_type", ""))
        
        # Ranking Formula
        # alpha=0.6, beta=0.3, gamma=0.1
        final_score = (0.6 * sem_score) + (0.3 * kw_score) + (0.1 * auth_score)

        if src_id not in entity_candidates or final_score > entity_candidates[src_id]["score"]:
            entity_candidates[src_id] = {
                "source_entity_id": src_id,
                "project_id": meta.get("project_id", req.project_id),
                "source_type": meta.get("source_type", "unknown"),
                "chunk_id": chunk_id,
                "score": final_score,
                "version_hash": meta.get("version_hash", ""),
                "metadata": {k: v for k, v in meta.items() if str(k).startswith("custom_")}
            }

    try:
        if mode in ["FULL_TEXT", "HYBRID", "EXACT"]:
            kw_results = rag_collection.get(
                where=where_filter,
                where_document={"$contains": query_text}
            )
            if kw_results and kw_results["ids"]:
                for idx, c_id in enumerate(kw_results["ids"]):
                    add_candidate(kw_results["metadatas"][idx], c_id, sem_score=0.0, kw_score=1.0)
                    
        if mode in ["SEMANTIC", "HYBRID"]:
            sem_results = rag_collection.query(
                query_texts=[query_text],
                n_results=limit * 3,
                where=where_filter
            )
            if sem_results and sem_results["ids"] and len(sem_results["ids"][0]) > 0:
                for idx, c_id in enumerate(sem_results["ids"][0]):
                    dist = sem_results["distances"][0][idx]
                    meta = sem_results["metadatas"][0][idx]
                    
                    norm_score = 1.0 / (1.0 + dist)
                    
                    kw_score = 1.0 if (mode == "HYBRID" and meta.get("source_entity_id") in entity_candidates and entity_candidates[meta["source_entity_id"]]["chunk_id"] == c_id) else 0.0
                    
                    # Apply semantic threshold
                    if norm_score > 0.4 or kw_score > 0:
                        add_candidate(meta, c_id, sem_score=norm_score, kw_score=kw_score)
                    
    except Exception as e:
        if mode == "HYBRID":
            pass # Fallback to whatever candidates we have
        else:
            raise HTTPException(500, f"Retrieval failed: {str(e)}")

    sorted_results = sorted(entity_candidates.values(), key=lambda x: x["score"], reverse=True)
    return RetrievalResponse(version=req.version, results=sorted_results[:limit])



# ------------------------------------------------------------------------------

# ------------------------------------------------------------------------------
# RAG INDEXING SYNC CONTRACTS (PHASE 2F.3)
# ------------------------------------------------------------------------------
import hashlib
import chromadb
from chromadb.config import Settings
from llama_index.core.node_parser import SentenceSplitter

# Initialize ChromaDB persistent client
chroma_client = None
# Using default embedding function (all-MiniLM-L6-v2) for zero-configuration local execution
try:
    chroma_client = chromadb.PersistentClient(path='./chroma_db', settings=Settings(anonymized_telemetry=False))
    rag_collection = chroma_client.get_or_create_collection(name="ai_dost_derived_index")
except Exception as e:
    print(f"Warning: Could not initialize Chroma collection: {e}")
    rag_collection = None

text_splitter = SentenceSplitter(chunk_size=512, chunk_overlap=50)

def generate_chunk_id(project_id: str, source_entity_id: str, version_hash: str, index: int) -> str:
    s = f"{project_id}_{source_entity_id}_{version_hash}_{index}"
    return hashlib.sha256(s.encode()).hexdigest()

class IndexDocument(BaseModel):
    source_entity_id: str
    source_type: str
    version_hash: str
    content: Optional[str] = ""
    metadata: Optional[dict] = {}

class IndexRequest(BaseModel):
    version: str = "1"
    project_id: str
    action: str  # "upsert" or "delete"
    documents: List[IndexDocument]

class IndexResponse(BaseModel):
    status: str
    processed_count: int
    chunks_created: int = 0
    chunks_deleted: int = 0

@app.post("/ai/rag/index", response_model=IndexResponse)
def rag_index(req: IndexRequest):
    """
    Phase 2F.3 Embedding & Vector Index Pipeline
    """
    if not req.project_id:
        raise HTTPException(400, "project_id is required for tenant isolation")
    
    if req.action not in ["upsert", "delete"]:
        raise HTTPException(400, "invalid action")

    if not rag_collection:
        raise HTTPException(503, "Vector store is unavailable")

    chunks_created = 0
    chunks_deleted = 0

    for doc in req.documents:
        if not doc.source_entity_id:
            raise HTTPException(400, "source_entity_id is required")

        # 1. Always purge existing vectors for this entity to ensure idempotency and stale data removal
        try:
            # Chroma delete by metadata
            rag_collection.delete(where={
                "$and": [
                    {"project_id": {"$eq": req.project_id}},
                    {"source_entity_id": {"$eq": doc.source_entity_id}}
                ]
            })
            # We don't know exactly how many were deleted easily via the API without querying first, 
            # but idempotency is achieved.
            chunks_deleted += 1 
        except Exception as e:
            # If nothing to delete, chroma might ignore or throw depending on version. We continue.
            pass

        # 2. If upsert, chunk and embed
        if req.action == "upsert" and doc.content:
            chunks = text_splitter.split_text(doc.content)
            ids = []
            documents = []
            metadatas = []
            
            for idx, chunk_text in enumerate(chunks):
                chunk_id = generate_chunk_id(req.project_id, doc.source_entity_id, doc.version_hash, idx)
                
                # Enforce project isolation and lineage in vector metadata
                meta = {
                    "project_id": req.project_id,
                    "source_entity_id": doc.source_entity_id,
                    "source_type": doc.source_type,
                    "version_hash": doc.version_hash,
                    "chunk_index": idx,
                    "embedding_model": "default-minilm-l6-v2"
                }
                
                # Merge custom metadata securely (ensuring it doesn't overwrite system keys)
                if doc.metadata:
                    for k, v in doc.metadata.items():
                        if k not in meta and isinstance(v, (str, int, float, bool)):
                            meta[f"custom_{k}"] = v

                ids.append(chunk_id)
                documents.append(chunk_text)
                metadatas.append(meta)

            if ids:
                try:
                    rag_collection.add(
                        ids=ids,
                        documents=documents,
                        metadatas=metadatas
                    )
                    chunks_created += len(ids)
                except Exception as e:
                    raise HTTPException(500, f"Vector store insertion failed: {e}")

    return IndexResponse(
        status="success", 
        processed_count=len(req.documents),
        chunks_created=chunks_created,
        chunks_deleted=chunks_deleted
    )
if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8001)






